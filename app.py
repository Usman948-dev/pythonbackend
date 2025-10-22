from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import camelot
import pandas as pd
import io
import os
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import easyocr
from pdf2image import convert_from_bytes
import numpy as np
from werkzeug.utils import secure_filename
import tempfile
import traceback

app = Flask(__name__)

# CORS configuration - CRITICAL for frontend to work
CORS(app, 
     resources={r"/api/*": {"origins": "*"}},
     allow_headers=["Content-Type"],
     expose_headers=["Content-Type"],
     supports_credentials=False)  # Enable CORS for frontend communication

# Configuration
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
UPLOAD_FOLDER = tempfile.gettempdir()
ALLOWED_EXTENSIONS = {'pdf'}

# --- Helper Functions and Data (Same as original) ---

UNIT_CONVERSIONS = {
    "p-no/p hour": 1.0,
    "p-no": 1.0,
    "hour": 1.0,
    "hours": 1.0,
    "meters": 1.0,
    "kilometers": 1000.0,
    "each": 1.0,
    "lot": 1.0,
    "kgs": 1.0,
    "tons": 1000.0,
    "liters": 1.0,
    "gallons": 3.785,
    "days": 8.0,
    "nos": 1.0,
    "pc": 1.0,
    "set": 1.0,
    "%": 100.0,
    "%0": 1000.0,
    "%o": 1000.0,
    "sqm": 1.0,
    "cum": 1.0,
    "lm": 1.0,
}

PDF_COLUMN_MAPPING_RULES = {
    "Sr No.": {
        "keywords": [
            "sr.no", "sr no.", "sr no", "seriel number", "srno", "s.no",
            "serial no", "serial number", "sr. no.", "sr.", "no.", "sno",
            "s. no", "srl no", "item no", "idx", "index", "chap no",
            "item no", "sl no", "s.l.no", "sl.no", "s.no"
        ],
        "include_header_cell_in_data": False
    },
    "Items Description": {
        "keywords": [
            "items", "item description", "items description", "item name", "items name",
            "item", "description", "item desc", "description of item", "desc", "particulars",
            "details", "material description", "scope of work", "work description", "narration",
            "description of item", "item details", "particuler of item", "particular"
        ],
        "include_header_cell_in_data": False
    },
    "Units": {
        "keywords": [
            "units", "unit", "uom", "measure", "unit of measure", "qty unit",
            "unit", "u.o.m", "unit of measurement"
        ],
        "include_header_cell_in_data": False
    },
    "Quantity": {
        "keywords": [
            "quantity", "quantities", "estimated quantity", "estimated qty", "qty",
            "est. qty", "est qty", "estimated quanity", "quanity", "total qty",
            "number", "nos", "volume", "amount", "no of items", "count", "number of items"
        ],
        "include_header_cell_in_data": False
    },
    "Govt Rate - Input": {
        "keywords": [
            "est", "est rates", "est rate", "estimated rates", "estimated rate",
            "rate", "rates", "market rate", "mkt rate", "mrkt rate", "market rates",
            "govt rate", "government rate", "official rate", "base rate", "approved rate",
            "tender rate", "schedule rate", "sch rate", "spec rate", "agreed rate",
            "rates (rs.)", "rates (r", "rates rs", "est price", "govt price"
        ],
        "include_header_cell_in_data": False
    },
    "Quoted Rate - Input": {
        "keywords": [
            "quoted rate", "quote rate", "quoted rates", "rates (rs.)", "offer rate",
            "bid rate", "client rate", "our rate", "contractor rate", "contract rates",
            "final rate", "negotiated rate", "amounts (rs.)", "amounts (r", "amounts rs",
            "amounts", "bid price", "offered price", "your rate"
        ],
        "include_header_cell_in_data": False
    },
}

FINAL_DISPLAY_EXCEL_COLUMN_ORDER = [
    "Sr No.", "Items Description", "Units", "Units No.", "Quantity",
    "Govt Rate - Input", "Govt Rate - Total", "Quoted Rate - Input", "Quoted Rate - Total"
]

CRITICAL_HEADER_COLUMNS = ["Sr No.", "Items Description", "Units", "Quantity"]
NON_CRITICAL_COLUMN_WEIGHT = 1
CRITICAL_COLUMN_WEIGHT = 5
MAX_HEADER_SCAN_ROWS = 10

COLUMN_WEIGHTS = {
    col: CRITICAL_COLUMN_WEIGHT if col in CRITICAL_HEADER_COLUMNS else NON_CRITICAL_COLUMN_WEIGHT
    for col in PDF_COLUMN_MAPPING_RULES.keys()
}


def convert_unit_to_number(unit_string):
    """Converts a unit string to a numerical value."""
    if not isinstance(unit_string, str):
        return 0.0

    unit_string_cleaned = unit_string.strip().lower()

    if "%0" in unit_string_cleaned or "%o" in unit_string_cleaned or "% 0" in unit_string_cleaned or "% o" in unit_string_cleaned:
        return 1000.0
    elif "%" in unit_string_cleaned:
        return 100.0

    return UNIT_CONVERSIONS.get(unit_string_cleaned, 1.0)


def calculate_total_rate(input_rate, quantity, units_no):
    """Calculates Total Rate: (Input Rate * Quantity) / Units No."""
    input_rate = pd.to_numeric(input_rate, errors='coerce').fillna(0.0)
    quantity = pd.to_numeric(quantity, errors='coerce').fillna(0.0)
    units_no = pd.to_numeric(units_no, errors='coerce').fillna(1.0)
    units_no = units_no.apply(lambda x: 1.0 if x == 0 else x)
    return (input_rate * quantity) / units_no


def recalculate_df_values(df):
    """Recalculates 'Units No.' and all 'Total Rate' columns."""
    df_copy = df.copy()

    if "Units" in df_copy.columns:
        new_units_no_from_units = df_copy["Units"].apply(convert_unit_to_number)
        df_copy["Units No."] = new_units_no_from_units

    df_copy["Units No."] = pd.to_numeric(df_copy["Units No."], errors='coerce').fillna(1.0)
    df_copy["Units No."] = df_copy["Units No."].apply(lambda x: 1.0 if x == 0 else x)

    for col in ["Quantity", "Govt Rate - Input", "Quoted Rate - Input"]:
        if col in df_copy.columns:
            df_copy[col] = pd.to_numeric(df_copy[col], errors='coerce').fillna(0.0)

    if all(col in df_copy.columns for col in ["Govt Rate - Input", "Quantity", "Units No."]):
        df_copy["Govt Rate - Total"] = calculate_total_rate(
            df_copy["Govt Rate - Input"], df_copy["Quantity"], df_copy["Units No."])

    if all(col in df_copy.columns for col in ["Quoted Rate - Input", "Quantity", "Units No."]):
        df_copy["Quoted Rate - Total"] = calculate_total_rate(
            df_copy["Quoted Rate - Input"], df_copy["Quantity"], df_copy["Units No."])

    return df_copy


def is_header_row(row_series, rules):
    """Checks if a given row is a header row."""
    row_values_cleaned = [str(cell).strip().lower() for cell in row_series.tolist()]
    match_count = 0
    for target_col_name, rule_keywords in rules.items():
        if any(keyword in cell_value for keyword in rule_keywords["keywords"] for cell_value in row_values_cleaned):
            match_count += 1
    return match_count >= 2


def fallback_header_detection(df):
    """Fallback method to identify columns based on content patterns."""
    temp_df = df.copy().astype(str).replace('', pd.NA)
    temp_df = temp_df.dropna(axis=1, how='all')
    temp_df = temp_df.iloc[:20, :]

    fallback_map = {}
    best_desc_score = -1
    best_desc_col_idx = None
    best_sr_no_score = -1
    best_sr_no_col_idx = None

    available_indices = list(range(len(temp_df.columns)))

    # Find best Items Description column
    for i, col in enumerate(temp_df.columns):
        series = temp_df[col].dropna()
        if not series.empty:
            avg_length = series.str.len().mean()
            alpha_ratio = series.str.replace(r'[^a-zA-Z\s]', '', regex=True).str.len().sum() / series.str.len().sum()
            score = avg_length * alpha_ratio
            if score > best_desc_score:
                best_desc_score = score
                best_desc_col_idx = i

    if best_desc_col_idx is not None:
        fallback_map["Items Description"] = best_desc_col_idx
        available_indices.remove(best_desc_col_idx)

    # Find best Sr No. column
    for i in available_indices:
        series = temp_df.iloc[:, i].dropna().reset_index(drop=True)
        if len(series) > 1:
            try:
                numeric_series = pd.to_numeric(series.str.replace(r'[^0-9.]', '', regex=True), errors='coerce')
                numeric_series_clean = numeric_series.dropna()
                if len(numeric_series_clean) > 1 and (numeric_series_clean.diff().dropna() > 0).all():
                    score = len(numeric_series_clean)
                    if score > best_sr_no_score:
                        best_sr_no_score = score
                        best_sr_no_col_idx = i
            except Exception:
                continue

    if best_sr_no_col_idx is not None:
        fallback_map["Sr No."] = best_sr_no_col_idx

    return fallback_map


def find_closest_x(x, x_coords):
    """Finds the index of the closest x coordinate."""
    return np.argmin(np.abs(x_coords - x))


def ocr_to_dataframe(results):
    """Parses OCR results to reconstruct a table-like DataFrame."""
    if not results:
        return pd.DataFrame()

    df_rows = []
    sorted_results = sorted(results, key=lambda r: (r[0][0][1], r[0][0][0]))

    rows = {}
    row_tol = 10
    for r in sorted_results:
        y_center = (r[0][0][1] + r[0][2][1]) / 2
        text = r[1]

        row_key = None
        for key in rows:
            if abs(key - y_center) < row_tol:
                row_key = key
                break

        if row_key is None:
            row_key = y_center
            rows[row_key] = []

        rows[row_key].append(r)

    all_x = [r[0][0][0] for r in sorted_results]
    unique_x = sorted(list(set(all_x)))

    column_centers = []
    if unique_x:
        column_centers.append(unique_x[0])
        for x_coord in unique_x[1:]:
            if x_coord - column_centers[-1] > 20:
                column_centers.append(x_coord)

    for row_key in sorted(rows.keys()):
        row_data = rows[row_key]
        row_cells = {}
        for r in row_data:
            x_center = (r[0][0][0] + r[0][1][0]) / 2
            text = r[1]
            col_idx = find_closest_x(x_center, np.array(column_centers))
            row_cells.setdefault(col_idx, []).append(text)

        row_dict = {}
        for col_idx, words in row_cells.items():
            row_dict[f'col_{col_idx}'] = ' '.join(words)
        df_rows.append(row_dict)

    if not df_rows:
        return pd.DataFrame()

    df = pd.DataFrame(df_rows).fillna("")

    new_cols = {}
    for col in df.columns:
        first_cell = df.loc[0, col]
        if ' ' in first_cell:
            new_cols[col] = first_cell.replace(' ', '_')
        else:
            new_cols[col] = first_cell

    df.rename(columns=new_cols, inplace=True)
    df = df.iloc[1:].reset_index(drop=True)

    return df


def extract_tables_with_ocr(pdf_bytes, pages):
    """Uses OCR to extract tables from a PDF."""
    try:
        images = convert_from_bytes(pdf_bytes)
    except Exception as e:
        return [], f"Error converting PDF to images: {str(e)}"

    reader = easyocr.Reader(['en'])
    all_dfs = []

    for i, image in enumerate(images):
        if pages != 'all' and str(i + 1) not in pages.split(','):
            continue

        try:
            results = reader.readtext(np.array(image))
            df = ocr_to_dataframe(results)
            if not df.empty:
                all_dfs.append(df)
        except Exception as e:
            continue

    return all_dfs, None


def process_pdf_extraction(pdf_bytes, pages_arg, table_areas_arg=None):
    """Main PDF extraction function."""
    pdf_stream = io.BytesIO(pdf_bytes)
    tables = []
    extraction_flavor = 'stream'
    extraction_method = None

    # Try manual coordinates if provided
    if table_areas_arg:
        try:
            tables = camelot.read_pdf(
                pdf_stream,
                pages=pages_arg,
                flavor='lattice',
                table_areas=table_areas_arg
            )
            extraction_method = "manual_coordinates"
        except Exception as e:
            return None, f"Extraction with coordinates failed: {str(e)}", None

    # Try automatic extraction
    if not table_areas_arg:
        try:
            tables = camelot.read_pdf(
                pdf_stream,
                pages=pages_arg,
                flavor=extraction_flavor,
                split_text=True,
                strip_text=' \n\r',
                row_tol=100
            )
            extraction_method = "stream"
        except Exception as stream_e:
            extraction_flavor = 'lattice'
            pdf_stream.seek(0)
            try:
                tables = camelot.read_pdf(
                    pdf_stream,
                    pages=pages_arg,
                    flavor=extraction_flavor,
                    split_text=True,
                    strip_text=' \n\r',
                    table_areas=None
                )
                extraction_method = "lattice"
            except Exception as lattice_e:
                # Try OCR as last resort
                pdf_stream.seek(0)
                extracted_dfs, ocr_error = extract_tables_with_ocr(pdf_bytes, pages_arg)
                if extracted_dfs:
                    return extracted_dfs, None, "ocr"
                else:
                    return None, "All extraction methods failed. Manual coordinates may be required.", None

    # Process extracted tables
    if len(tables) == 0 or all(t.df.empty for t in tables):
        pdf_stream.seek(0)
        extracted_dfs, ocr_error = extract_tables_with_ocr(pdf_bytes, pages_arg)
        if extracted_dfs:
            return extracted_dfs, None, "ocr"
        else:
            return None, "No tables found in PDF", None

    # Convert Camelot tables to DataFrames
    extracted_dfs = [t.df for t in sorted(tables, key=lambda t: t.page)]
    return extracted_dfs, None, extraction_method


def process_dataframes(extracted_dfs):
    """Process extracted DataFrames with header detection and column mapping."""
    if not extracted_dfs:
        return None, "No dataframes to process"

    df_first_page = extracted_dfs[0].copy()
    df_first_page.columns = [str(col).strip() for col in df_first_page.columns]

    # Header detection
    best_header_row_idx = -1
    max_weighted_matches = 0
    header_column_map = {}
    keyword_based_success = False

    for row_idx in range(min(MAX_HEADER_SCAN_ROWS, df_first_page.shape[0])):
        current_row_values_cleaned = [str(cell).strip().lower() for cell in df_first_page.iloc[row_idx].tolist()]
        current_weighted_matches = 0
        temp_header_col_map = {}

        for target_col_name, rules in PDF_COLUMN_MAPPING_RULES.items():
            for i, header_cell_content in enumerate(current_row_values_cleaned):
                if any(keyword in header_cell_content for keyword in rules["keywords"]):
                    temp_header_col_map[target_col_name] = i
                    current_weighted_matches += COLUMN_WEIGHTS.get(target_col_name, NON_CRITICAL_COLUMN_WEIGHT)
                    break

        if current_weighted_matches > max_weighted_matches:
            max_weighted_matches = current_weighted_matches
            best_header_row_idx = row_idx
            header_column_map = temp_header_col_map.copy()

        if all(col in header_column_map for col in CRITICAL_HEADER_COLUMNS):
            keyword_based_success = True
            break

    # Fallback detection if keywords failed
    if not keyword_based_success:
        header_column_map = fallback_header_detection(df_first_page)
        if header_column_map:
            best_header_row_idx = -1

    if not header_column_map:
        return None, "Could not identify header row"

    # Process all DataFrames
    all_processed_dfs = []

    for idx, df in enumerate(extracted_dfs):
        df_copy = df.copy()
        df_copy.columns = [str(col).strip() for col in df_copy.columns]

        start_data_row = best_header_row_idx + 1 if best_header_row_idx != -1 else 0

        # Skip header on subsequent pages
        if idx > 0 and not df_copy.empty and is_header_row(df_copy.iloc[0], PDF_COLUMN_MAPPING_RULES):
            start_data_row = 1

        current_processed_df_data = {}

        for target_col_name in FINAL_DISPLAY_EXCEL_COLUMN_ORDER:
            original_col_index = header_column_map.get(target_col_name)

            if original_col_index is not None and original_col_index < df_copy.shape[1]:
                col_data = df_copy.iloc[start_data_row:, original_col_index].reset_index(drop=True)
                current_processed_df_data[target_col_name] = col_data
            else:
                series_length = df_copy.shape[0] - start_data_row
                current_processed_df_data[target_col_name] = pd.Series(dtype='object', index=range(series_length))

        extracted_flat_df = pd.DataFrame(current_processed_df_data)

        # Initialize calculated columns
        for col in ["Quantity", "Govt Rate - Input", "Quoted Rate - Input"]:
            if col in extracted_flat_df.columns:
                extracted_flat_df[col] = pd.to_numeric(extracted_flat_df[col], errors='coerce').fillna(0.0)

        for col in ["Units No.", "Govt Rate - Total", "Quoted Rate - Total"]:
            if col in extracted_flat_df.columns:
                extracted_flat_df[col] = 0.0

        extracted_flat_df = recalculate_df_values(extracted_flat_df)
        all_processed_dfs.append(extracted_flat_df)

    combined_df = pd.concat(all_processed_dfs, ignore_index=True)
    return combined_df, None


def generate_excel(df, name=""):
    """Generates Excel file from DataFrame with formulas."""
    excel_buffer = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    sheet_name = "Combined Tables"
    ws = wb.create_sheet(title=sheet_name)

    col_to_excel_letter = {col_name: get_column_letter(idx + 1)
                           for idx, col_name in enumerate(FINAL_DISPLAY_EXCEL_COLUMN_ORDER)}

    initial_data_row = 4

    for _ in range(initial_data_row - 1):
        ws.append([])

    ws.append(FINAL_DISPLAY_EXCEL_COLUMN_ORDER)

    header_row_obj = ws[initial_data_row]
    for cell in header_row_obj:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

    ws.freeze_panes = get_column_letter(1) + str(initial_data_row + 1)

    # Add data rows with formulas
    for r_idx, row_data_series in df.iterrows():
        excel_row_num = r_idx + initial_data_row + 1
        row_values_for_excel = []

        for col_name in FINAL_DISPLAY_EXCEL_COLUMN_ORDER:
            if col_name in ["Govt Rate - Total", "Quoted Rate - Total"]:
                input_rate_col_name = "Govt Rate - Input" if col_name == "Govt Rate - Total" else "Quoted Rate - Input"
                input_rate_col_letter = col_to_excel_letter.get(input_rate_col_name)
                quantity_col_letter = col_to_excel_letter.get("Quantity")
                units_no_col_letter = col_to_excel_letter.get("Units No.")

                if all([input_rate_col_letter, quantity_col_letter, units_no_col_letter]):
                    formula_str = (
                        f"=(IFERROR(VALUE({input_rate_col_letter}{excel_row_num}),0)*IFERROR(VALUE({quantity_col_letter}{excel_row_num}),0))/"
                        f"(IF(OR(ISBLANK({units_no_col_letter}{excel_row_num}),IFERROR(VALUE({units_no_col_letter}{excel_row_num}),1)=0),1,IFERROR(VALUE({units_no_col_letter}{excel_row_num}),1)))"
                    )
                    row_values_for_excel.append(formula_str)
                else:
                    row_values_for_excel.append(None)
            else:
                cell_value = row_data_series.get(col_name)
                row_values_for_excel.append(None if pd.isna(cell_value) else str(cell_value).strip())

        ws.append(row_values_for_excel)

    # Set column widths
    for col_idx, column_header in enumerate(FINAL_DISPLAY_EXCEL_COLUMN_ORDER):
        col_letter = get_column_letter(col_idx + 1)
        if column_header == "Items Description":
            ws.column_dimensions[col_letter].width = 70
            for r in range(initial_data_row + 1, ws.max_row + 1):
                ws.cell(row=r, column=col_idx + 1).alignment = Alignment(wrap_text=True, vertical='top')
        else:
            ws.column_dimensions[col_letter].width = 12
            for r in range(initial_data_row + 1, ws.max_row + 1):
                ws.cell(row=r, column=col_idx + 1).alignment = Alignment(wrap_text=True, vertical='top')

    # Add borders
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=initial_data_row, max_row=ws.max_row, 
                           min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    current_data_end_row = ws.max_row

    # Add grand totals
    govt_total_col_letter = col_to_excel_letter.get("Govt Rate - Total")
    quoted_total_col_letter = col_to_excel_letter.get("Quoted Rate - Total")

    if govt_total_col_letter:
        grand_total_label_row = [None] * len(FINAL_DISPLAY_EXCEL_COLUMN_ORDER)
        grand_total_label_row[0] = "Grand Total (Govt)"
        ws.append(grand_total_label_row)
        row_num_govt = ws.max_row

        cell_govt = ws.cell(row=row_num_govt, 
                           column=FINAL_DISPLAY_EXCEL_COLUMN_ORDER.index("Govt Rate - Total") + 1)
        cell_govt.value = f"=SUM({govt_total_col_letter}{initial_data_row + 1}:{govt_total_col_letter}{current_data_end_row})"
        for cell in ws[row_num_govt]:
            cell.font = Font(bold=True)
            cell.border = thin_border

    if quoted_total_col_letter:
        grand_total_label_row = [None] * len(FINAL_DISPLAY_EXCEL_COLUMN_ORDER)
        grand_total_label_row[0] = "Grand Total (Quoted)"
        ws.append(grand_total_label_row)
        row_num_quoted = ws.max_row

        cell_quoted = ws.cell(row=row_num_quoted, 
                             column=FINAL_DISPLAY_EXCEL_COLUMN_ORDER.index("Quoted Rate - Total") + 1)
        cell_quoted.value = f"=SUM({quoted_total_col_letter}{initial_data_row + 1}:{quoted_total_col_letter}{current_data_end_row})"
        for cell in ws[row_num_quoted]:
            cell.font = Font(bold=True)
            cell.border = thin_border

    # Add name
    if name:
        ws['A1'] = name
        ws['A1'].font = Font(bold=True)
        ws['A1'].border = thin_border

    # Add rate below formula
    if govt_total_col_letter and quoted_total_col_letter and 'row_num_govt' in locals() and 'row_num_quoted' in locals():
        ws[f'{govt_total_col_letter}1'] = 'Rate Below From Govt Rate'
        ws[f'{govt_total_col_letter}1'].font = Font(bold=True)
        ws[f'{govt_total_col_letter}1'].alignment = Alignment(horizontal='right')
        ws[f'{govt_total_col_letter}1'].border = thin_border
        ws[f'{govt_total_col_letter}1'].fill = PatternFill(start_color="FFF2CC", 
                                                           end_color="FFF2CC", fill_type="solid")

        formula_cell_loc = f'{quoted_total_col_letter}1'
        formula_summary_str = (
            f'=IFERROR(1-(IFERROR(VALUE({quoted_total_col_letter}{row_num_quoted}),0)/'
            f'IF(IFERROR(VALUE({govt_total_col_letter}{row_num_govt}),0)=0,1,IFERROR(VALUE({govt_total_col_letter}{row_num_govt}),0))),0)'
        )
        ws[formula_cell_loc] = formula_summary_str
        ws[formula_cell_loc].number_format = '0.00%'
        ws[formula_cell_loc].font = Font(bold=True)
        ws[formula_cell_loc].border = thin_border
        ws[formula_cell_loc].fill = PatternFill(start_color="D9EAD3", 
                                               end_color="D9EAD3", fill_type="solid")

    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


# --- API Routes ---

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({"status": "healthy", "message": "PDF Extractor API is running"}), 200


@app.route('/api/extract', methods=['POST'])
def extract_pdf():
    """
    Extract tables from PDF.
    Form data:
    - file: PDF file
    - pages: Page numbers (e.g., "1,3-5" or empty for all)
    - table_areas: Manual coordinates (optional)
    - name: Name for Excel sheet
    """
    try:
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400

        if not file.filename.lower().endswith('.pdf'):
            return jsonify({"error": "File must be a PDF"}), 400

        # Get parameters
        pages_input = request.form.get('pages', '')
        pages_arg = 'all' if pages_input.strip() == "" else pages_input
        
        table_areas_input = request.form.get('table_areas', '')
        table_areas_arg = None
        if table_areas_input:
            table_areas_arg = [float(x) for x in re.split(r'[,\s]+', table_areas_input.strip())]
            table_areas_arg = [' '.join(map(str, table_areas_arg[i:i + 4])) 
                             for i in range(0, len(table_areas_arg), 4)]

        name = request.form.get('name', 'Name:')

        # Read PDF bytes
        pdf_bytes = file.read()

        # Extract tables
        extracted_dfs, error, extraction_method = process_pdf_extraction(
            pdf_bytes, pages_arg, table_areas_arg
        )

        if error:
            return jsonify({
                "error": error,
                "requires_manual_coordinates": "coordinates" in error.lower()
            }), 400

        # Process dataframes
        combined_df, process_error = process_dataframes(extracted_dfs)

        if process_error:
            return jsonify({"error": process_error}), 400

        # Calculate summary
        total_govt = combined_df["Govt Rate - Total"].sum() if "Govt Rate - Total" in combined_df.columns else 0
        total_quoted = combined_df["Quoted Rate - Total"].sum() if "Quoted Rate - Total" in combined_df.columns else 0
        rate_below = (1 - (total_quoted / total_govt)) if total_govt > 0 else 0

        # Convert DataFrame to dict for JSON response
        data_dict = combined_df.to_dict(orient='records')

        return jsonify({
            "success": True,
            "extraction_method": extraction_method,
            "data": data_dict,
            "summary": {
                "name": name,
                "grand_total_govt": float(total_govt),
                "grand_total_quoted": float(total_quoted),
                "rate_below_govt": float(rate_below)
            },
            "columns": FINAL_DISPLAY_EXCEL_COLUMN_ORDER
        }), 200

    except Exception as e:
        return jsonify({
            "error": f"Server error: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500


@app.route('/api/recalculate', methods=['POST'])
def recalculate_data():
    """
    Recalculate totals based on edited data.
    JSON body:
    - data: Array of row objects
    - name: Name for summary
    """
    try:
        json_data = request.get_json()
        
        if not json_data or 'data' not in json_data:
            return jsonify({"error": "No data provided"}), 400

        data = json_data['data']
        name = json_data.get('name', 'Name:')

        # Convert to DataFrame
        df = pd.DataFrame(data)

        # Ensure all required columns exist
        for col in FINAL_DISPLAY_EXCEL_COLUMN_ORDER:
            if col not in df.columns:
                df[col] = None

        # Recalculate
        df = recalculate_df_values(df)

        # Calculate summary
        total_govt = df["Govt Rate - Total"].sum() if "Govt Rate - Total" in df.columns else 0
        total_quoted = df["Quoted Rate - Total"].sum() if "Quoted Rate - Total" in df.columns else 0
        rate_below = (1 - (total_quoted / total_govt)) if total_govt > 0 else 0

        # Convert back to dict
        data_dict = df.to_dict(orient='records')

        return jsonify({
            "success": True,
            "data": data_dict,
            "summary": {
                "name": name,
                "grand_total_govt": float(total_govt),
                "grand_total_quoted": float(total_quoted),
                "rate_below_govt": float(rate_below)
            }
        }), 200

    except Exception as e:
        return jsonify({
            "error": f"Recalculation error: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500


@app.route('/api/download-excel', methods=['POST'])
def download_excel():
    """
    Generate and download Excel file.
    JSON body:
    - data: Array of row objects
    - name: Name for Excel sheet
    """
    try:
        json_data = request.get_json()
        
        if not json_data or 'data' not in json_data:
            return jsonify({"error": "No data provided"}), 400

        data = json_data['data']
        name = json_data.get('name', 'Name:')

        # Convert to DataFrame
        df = pd.DataFrame(data)

        # Ensure all required columns exist
        for col in FINAL_DISPLAY_EXCEL_COLUMN_ORDER:
            if col not in df.columns:
                df[col] = None

        # Recalculate to ensure consistency
        df = recalculate_df_values(df)

        # Generate Excel
        excel_buffer = generate_excel(df, name)

        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='edited_tables_with_formulas.xlsx'
        )

    except Exception as e:
        return jsonify({
            "error": f"Excel generation error: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500


@app.route('/api/unit-conversions', methods=['GET'])
def get_unit_conversions():
    """Get available unit conversions."""
    return jsonify({
        "unit_conversions": UNIT_CONVERSIONS,
        "column_rules": {
            key: {"keywords": val["keywords"]} 
            for key, val in PDF_COLUMN_MAPPING_RULES.items()
        }
    }), 200


@app.errorhandler(413)
def too_large(e):
    return jsonify({"error": "File too large. Maximum size is 50MB"}), 413


@app.errorhandler(500)
def internal_error(e):
    return jsonify({"error": "Internal server error", "details": str(e)}), 500

@app.route('/', methods=['GET'])
def home():
    return jsonify({
        "status": "running",
        "message": "PDF Table Extractor API",
        "endpoints": {
            "/api/health": "GET - Health check",
            "/api/extract": "POST - Extract tables from PDF",
            "/api/recalculate": "POST - Recalculate totals",
            "/api/download-excel": "POST - Generate Excel file",
            "/api/unit-conversions": "GET - Get unit conversion rules"
        }
    }), 200


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
